from openpyxl import load_workbook

#  Excel 
wb = load_workbook("file_05/OpenPI.xlsx")

# SheetDFHI
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        d_val = ws.cell(row=row, column=4).value  # D(4)
        f_val = ws.cell(row=row, column=6).value  # F(6)
        h_val = ws.cell(row=row, column=8).value  # H(8)

        if None not in (d_val, f_val, h_val):
            avg = (d_val + f_val + h_val) / 3
            ws.cell(row=row, column=9, value=avg)  # I(9)

# SheetR
merged_ws = wb.create_sheet(title="Merged_Data")

# Sheet
for col_idx, sheet_name in enumerate(wb.sheetnames[:-1], start=1):  # Merged_Data
    merged_ws.cell(row=1, column=col_idx, value=sheet_name)

# SheetISheet
max_rows = max([wb[sheet].max_row for sheet in wb.sheetnames[:-1]])  # 

for col_idx, sheet_name in enumerate(wb.sheetnames[:-1], start=1):
    ws = wb[sheet_name]
    for row_idx in range(1, max_rows + 1):
        # I(column=9)
        if row_idx <= ws.max_row:
            cell_value = ws.cell(row=row_idx, column=9).value
            merged_ws.cell(row=row_idx + 1, column=col_idx, value=cell_value)
        else:
            merged_ws.cell(row=row_idx + 1, column=col_idx, value=None)

# 
wb.save("merge_data/OpenPI.xlsx")